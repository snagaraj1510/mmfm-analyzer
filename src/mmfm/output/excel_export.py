"""
Excel report export for MMFM Analyzer.

Generates a multi-sheet Excel workbook with financial analysis results.
Uses openpyxl for formatting and data population.
"""

from __future__ import annotations

import math
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

from mmfm.engine.core_metrics import NPVResult, IRRResult, PaybackResult
from mmfm.engine.projections import CashFlowProjection
from mmfm.engine.scenarios import ScenarioComparison
from mmfm.engine.sensitivity import SensitivityResult
from mmfm.engine.monte_carlo import MonteCarloResult


# ── Style helpers ────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

WHITE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
REGULAR_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def _set_col_widths(ws, widths: dict) -> None:
    """Set column widths. Keys are column letters or numbers."""
    for col, width in widths.items():
        if isinstance(col, int):
            col = get_column_letter(col)
        ws.column_dimensions[col].width = width


def _write_header_row(ws, row: int, headers: list[str], start_col: int = 1) -> None:
    for i, header in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _safe(value: float) -> Optional[float]:
    """Return None for NaN/Inf (Excel can't handle these)."""
    if math.isnan(value) or math.isinf(value):
        return None
    return value


# ── Sheet builders ───────────────────────────────────────────────────────────

def _build_summary_sheet(
    ws,
    npv: NPVResult,
    irr: IRRResult,
    payback: PaybackResult,
    source_file: str,
    currency: str,
) -> None:
    ws.title = "Summary"

    # Title
    ws["A1"] = "MMFM Financial Analysis Report"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = REGULAR_FONT
    ws["A3"] = f"Source: {source_file}"
    ws["A3"].font = REGULAR_FONT

    # Core metrics table
    ws["A5"] = "Core Financial Metrics"
    ws["A5"].font = WHITE_FONT
    ws["A5"].fill = SUBHEADER_FILL

    headers = ["Metric", "Value", "Status"]
    _write_header_row(ws, 6, headers)

    rows = [
        ("NPV", _safe(npv.value), "POSITIVE" if npv.is_positive else "NEGATIVE"),
        ("IRR", _safe(irr.value) if irr.converged else None, "OK" if irr.converged else "N/A"),
        ("Payback Period (years)", _safe(payback.years) if payback.reached else None,
         "REACHED" if payback.reached else "NOT REACHED"),
        ("Discount Rate", npv.discount_rate, "—"),
    ]

    for r_idx, (metric, value, status) in enumerate(rows):
        row = 7 + r_idx
        ws.cell(row=row, column=1, value=metric).font = REGULAR_FONT
        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.font = REGULAR_FONT
        val_cell.alignment = RIGHT
        status_cell = ws.cell(row=row, column=3, value=status)
        status_cell.alignment = CENTER

        # Color status cell
        if status in ("POSITIVE", "REACHED", "OK"):
            status_cell.fill = GREEN_FILL
        elif status in ("NEGATIVE", "NOT REACHED"):
            status_cell.fill = RED_FILL

        if r_idx % 2 == 1:
            for c in range(1, 4):
                ws.cell(row=row, column=c).fill = ALT_ROW_FILL

    _set_col_widths(ws, {"A": 30, "B": 18, "C": 18})


def _build_projection_sheet(ws, projection: CashFlowProjection) -> None:
    ws.title = "Cash Flow Projection"
    currency = projection.base_currency

    headers = ["Year", f"Revenue ({currency})", f"Capex ({currency})",
               f"Opex ({currency})", f"NOI ({currency})", f"FCF ({currency})",
               f"Cumulative ({currency})", "Occupancy Rate", "Op Margin"]
    _write_header_row(ws, 1, headers)

    for r_idx, row in enumerate(projection.years):
        r = r_idx + 2
        fill = ALT_ROW_FILL if r_idx % 2 == 1 else None

        values = [
            row.year,
            _safe(row.revenue),
            _safe(row.capex),
            _safe(row.opex),
            _safe(row.net_operating_income),
            _safe(row.free_cash_flow),
            _safe(row.cumulative_cash_flow),
            _safe(row.occupancy_rate),
            _safe(row.operating_margin),
        ]
        for c_idx, val in enumerate(values):
            cell = ws.cell(row=r, column=c_idx + 1, value=val)
            cell.font = REGULAR_FONT
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill

            # Format numbers
            if c_idx in (1, 2, 3, 4, 5, 6):
                cell.number_format = '#,##0'
            elif c_idx in (7, 8):
                cell.number_format = '0.0%'

            # Color FCF and cumulative
            if c_idx in (5, 6) and val is not None:
                if val < 0:
                    cell.fill = RED_FILL
                elif val > 0:
                    cell.fill = GREEN_FILL

    _set_col_widths(ws, {i: 16 for i in range(1, 10)})
    ws.column_dimensions["A"].width = 8


def _build_scenario_sheet(ws, comparison: ScenarioComparison) -> None:
    ws.title = "Scenario Comparison"

    headers = ["Scenario", "NPV", "IRR", "Payback (yrs)", "Min DSCR", "Avg Op Margin", "Description"]
    _write_header_row(ws, 1, headers)

    ranking = comparison.npv_ranking()
    for r_idx, scenario_name in enumerate(ranking):
        result = comparison.results[scenario_name]
        r = r_idx + 2

        row_data = [
            result.scenario.name.upper(),
            _safe(result.npv.value),
            _safe(result.irr.value) if result.irr.converged else None,
            _safe(result.payback.years) if result.payback.reached else None,
            _safe(result.dscr.min_dscr),
            _safe(result.operating_margin.average),
            result.scenario.description,
        ]

        # Best scenario gets green, worst gets red
        row_fill = None
        if r_idx == 0:
            row_fill = GREEN_FILL
        elif r_idx == len(ranking) - 1:
            row_fill = RED_FILL

        for c_idx, val in enumerate(row_data):
            cell = ws.cell(row=r, column=c_idx + 1, value=val)
            cell.font = REGULAR_FONT
            cell.border = THIN_BORDER
            if row_fill:
                cell.fill = row_fill
            if c_idx == 1:
                cell.number_format = '#,##0'
            elif c_idx in (2, 4, 5):
                cell.number_format = '0.0%'
            elif c_idx == 3:
                cell.number_format = '0.0'

    _set_col_widths(ws, {1: 14, 2: 16, 3: 10, 4: 14, 5: 12, 6: 16, 7: 45})


def _build_sensitivity_sheet(ws, sensitivity: SensitivityResult) -> None:
    ws.title = "Sensitivity Analysis"

    ws["A1"] = "Tornado Analysis — NPV Sensitivity"
    ws["A1"].font = BOLD_FONT
    ws["A2"] = f"Base NPV: {sensitivity.base_npv:,.2f}"

    headers = ["Variable", "Base Value", "Low NPV", "High NPV", "NPV Swing", "% Swing vs Base"]
    _write_header_row(ws, 4, headers)

    for r_idx, var in enumerate(sensitivity.tornado_order()):
        r = r_idx + 5
        base_npv = sensitivity.base_npv
        pct_swing = (var.npv_swing / abs(base_npv) * 100) if base_npv != 0 else 0.0

        row_data = [
            var.label,
            _safe(var.base_value),
            _safe(var.npv_at_low),
            _safe(var.npv_at_high),
            _safe(var.npv_swing),
            pct_swing,
        ]
        fill = ALT_ROW_FILL if r_idx % 2 == 1 else None
        for c_idx, val in enumerate(row_data):
            cell = ws.cell(row=r, column=c_idx + 1, value=val)
            cell.font = REGULAR_FONT
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
            if c_idx in (2, 3, 4):
                cell.number_format = '#,##0'
            elif c_idx == 5:
                cell.number_format = '0.0"%"'

    _set_col_widths(ws, {1: 28, 2: 14, 3: 16, 4: 16, 5: 14, 6: 18})


def _build_monte_carlo_sheet(ws, mc: MonteCarloResult) -> None:
    ws.title = "Monte Carlo"

    ws["A1"] = "Monte Carlo Simulation Results"
    ws["A1"].font = BOLD_FONT
    ws["A2"] = f"Iterations: {mc.iterations:,}  |  Seed: {mc.seed or 'random'}"

    headers = ["Metric", "Value"]
    _write_header_row(ws, 4, headers)

    rows = [
        ("NPV — P10 (pessimistic)", _safe(mc.npv_p10)),
        ("NPV — P50 (median)", _safe(mc.npv_p50)),
        ("NPV — P90 (optimistic)", _safe(mc.npv_p90)),
        ("NPV — Mean", _safe(mc.npv_mean)),
        ("NPV — Std Dev", _safe(mc.npv_std)),
        ("Probability of Positive NPV", mc.prob_positive_npv),
        ("IRR — P10", _safe(mc.irr_p10) if mc.irr_values else None),
        ("IRR — P50", _safe(mc.irr_p50) if mc.irr_values else None),
        ("IRR — P90", _safe(mc.irr_p90) if mc.irr_values else None),
        (f"Probability DSCR < {mc.dscr_threshold:.1f}x", mc.prob_dscr_below_threshold),
    ]

    for r_idx, (metric, value) in enumerate(rows):
        r = r_idx + 5
        fill = ALT_ROW_FILL if r_idx % 2 == 1 else None
        cell_m = ws.cell(row=r, column=1, value=metric)
        cell_v = ws.cell(row=r, column=2, value=value)
        cell_m.font = REGULAR_FONT
        cell_v.font = REGULAR_FONT
        cell_v.alignment = RIGHT
        if fill:
            cell_m.fill = fill
            cell_v.fill = fill
        if metric.startswith("NPV"):
            cell_v.number_format = '#,##0'
        elif "Probability" in metric or "IRR" in metric:
            cell_v.number_format = '0.0%'

    # Input correlations
    if mc.input_npv_correlations:
        ws.cell(row=len(rows) + 7, column=1, value="Input-NPV Correlations").font = BOLD_FONT
        _write_header_row(ws, len(rows) + 8, ["Input Variable", "Pearson r (with NPV)"])
        for r_idx, (var, corr) in enumerate(mc.input_npv_correlations.items()):
            r = len(rows) + 9 + r_idx
            ws.cell(row=r, column=1, value=var).font = REGULAR_FONT
            cell = ws.cell(row=r, column=2, value=_safe(corr))
            cell.font = REGULAR_FONT
            cell.number_format = '0.000'
            cell.alignment = RIGHT

    _set_col_widths(ws, {1: 35, 2: 18})


# ── Public API ───────────────────────────────────────────────────────────────

def export_excel(
    output_path: "Path | str",
    npv: NPVResult,
    irr: IRRResult,
    payback: PaybackResult,
    projection: CashFlowProjection,
    source_file: str = "",
    comparison: Optional[ScenarioComparison] = None,
    sensitivity: Optional[SensitivityResult] = None,
    monte_carlo: Optional[MonteCarloResult] = None,
) -> None:
    """
    Export analysis results to a formatted Excel workbook.

    Always includes: Summary, Cash Flow Projection sheets.
    Optional: Scenario Comparison, Sensitivity Analysis, Monte Carlo.

    Args:
        output_path: Where to save the .xlsx file
        npv: NPV result from core_metrics
        irr: IRR result from core_metrics
        payback: Payback result from core_metrics
        projection: Cash flow projection
        source_file: Source model filename (for report header)
        comparison: Optional scenario comparison results
        sensitivity: Optional sensitivity analysis results
        monte_carlo: Optional Monte Carlo simulation results
    """
    wb = Workbook()
    # Remove default sheet
    default_ws = wb.active
    wb.remove(default_ws)

    _build_summary_sheet(
        wb.create_sheet("Summary"), npv, irr, payback,
        source_file=source_file, currency=projection.base_currency,
    )
    _build_projection_sheet(wb.create_sheet("Cash Flow Projection"), projection)

    if comparison:
        _build_scenario_sheet(wb.create_sheet("Scenario Comparison"), comparison)

    if sensitivity:
        _build_sensitivity_sheet(wb.create_sheet("Sensitivity Analysis"), sensitivity)

    if monte_carlo:
        _build_monte_carlo_sheet(wb.create_sheet("Monte Carlo"), monte_carlo)

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
