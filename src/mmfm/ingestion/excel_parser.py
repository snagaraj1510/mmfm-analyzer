"""
Excel financial model parser.

Parses .xlsx and .xlsm files, extracts named ranges, and validates
against registered schemas. Returns structured data for the financial engine.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd

from mmfm.ingestion.schema_validator import validate_sheets, load_schema, ValidationResult
from mmfm.config import SCHEMAS_DIR


# Maps schema names to the resource type they represent
SCHEMA_ALIASES = {
    "revenue_schema": "revenue",
    "capex_schema": "capex",
    "opex_schema": "opex",
    "assumptions_schema": "assumptions",
}


@dataclass
class ParsedExcelModel:
    """Structured representation of a parsed Excel financial model."""
    source_file: Path
    sheets: dict[str, pd.DataFrame] = field(default_factory=dict)
    named_ranges: dict[str, float | str] = field(default_factory=dict)
    detected_schema: Optional[str] = None
    validation_result: Optional[ValidationResult] = None
    base_currency: str = "USD"
    metadata: dict = field(default_factory=dict)


def _read_named_ranges(wb: openpyxl.Workbook) -> dict[str, float | str]:
    """Extract named ranges (scalar values) from a workbook."""
    named_ranges: dict[str, float | str] = {}
    for named_range in wb.defined_names.values():
        try:
            destinations = list(named_range.destinations)
            if len(destinations) == 1:
                sheet_name, cell_ref = destinations[0]
                ws = wb[sheet_name]
                # Strip $ signs for cell reference
                clean_ref = cell_ref.replace("$", "")
                cell = ws[clean_ref]
                if cell.value is not None:
                    named_ranges[named_range.name] = cell.value
        except Exception:
            # Named ranges can be complex (multi-cell, cross-sheet) — skip those
            pass
    return named_ranges


def _detect_base_currency(sheets: dict[str, pd.DataFrame], named_ranges: dict) -> str:
    """Attempt to detect the base currency from named ranges or cell content."""
    # Check named ranges first
    for key in named_ranges:
        if "currency" in key.lower() or "base_currency" in key.lower():
            val = str(named_ranges[key]).strip().upper()
            if val in ("USD", "KES", "TZS", "MZN", "EUR"):
                return val

    # Check for currency symbols in first 50 cells of each sheet
    currency_map = {"$": "USD", "KES": "KES", "TZS": "TZS", "MZN": "MZN", "€": "EUR"}
    for df in sheets.values():
        sample = df.head(10).astype(str).values.flatten()
        for cell_val in sample:
            for symbol, currency in currency_map.items():
                if symbol in cell_val:
                    return currency

    return "USD"  # Default


def _detect_schema(sheets: dict[str, pd.DataFrame]) -> Optional[str]:
    """Heuristically detect which schema matches based on sheet names."""
    sheet_names_lower = {name.lower() for name in sheets}

    if "revenue projections" in sheet_names_lower or "revenue" in sheet_names_lower:
        return "revenue_schema"
    if "capital expenditure" in sheet_names_lower or "capex" in sheet_names_lower:
        return "capex_schema"
    if "operating expenditure" in sheet_names_lower or "opex" in sheet_names_lower:
        return "opex_schema"
    if "assumptions" in sheet_names_lower:
        return "assumptions_schema"

    return None


def parse_excel(
    file_path: Path | str,
    schema_name: Optional[str] = None,
    validate: bool = True,
) -> ParsedExcelModel:
    """
    Parse an Excel financial model file.

    Args:
        file_path: Path to .xlsx or .xlsm file
        schema_name: Schema to validate against (auto-detected if None)
        validate: Whether to run schema validation

    Returns:
        ParsedExcelModel with sheets, named ranges, and validation result

    Raises:
        FileNotFoundError: If the file does not exist
        ValueError: If the file is not a supported format
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        raise ValueError(f"Unsupported file format: {path.suffix}. Use .xlsx or .xlsm")

    wb = openpyxl.load_workbook(path, data_only=True)

    # Read all sheets into DataFrames
    sheets: dict[str, pd.DataFrame] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = list(ws.values)
        if not data:
            continue
        # First row as header
        headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(data[0])]
        rows = data[1:]
        df = pd.DataFrame(rows, columns=headers)
        # Drop fully-empty rows
        df = df.dropna(how="all")
        sheets[sheet_name] = df

    named_ranges = _read_named_ranges(wb)
    detected_schema = schema_name or _detect_schema(sheets)
    base_currency = _detect_base_currency(sheets, named_ranges)

    model = ParsedExcelModel(
        source_file=path,
        sheets=sheets,
        named_ranges=named_ranges,
        detected_schema=detected_schema,
        base_currency=base_currency,
        metadata={
            "sheet_names": list(sheets.keys()),
            "row_counts": {name: len(df) for name, df in sheets.items()},
        },
    )

    if validate and detected_schema:
        schema_path = SCHEMAS_DIR / f"{detected_schema}.yaml"
        if schema_path.exists():
            schema = load_schema(detected_schema)
            model.validation_result = validate_sheets(sheets, schema)
        else:
            # Schema file doesn't exist yet — skip validation
            pass

    return model
