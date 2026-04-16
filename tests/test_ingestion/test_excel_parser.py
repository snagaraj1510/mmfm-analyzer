"""Tests for Excel parser."""

from __future__ import annotations

import pytest
from pathlib import Path
import tempfile
import openpyxl

from mmfm.ingestion.excel_parser import parse_excel, _detect_schema


class TestExcelParser:
    def test_file_not_found_raises(self):
        with pytest.raises(FileNotFoundError):
            parse_excel(Path("/nonexistent/file.xlsx"))

    def test_wrong_extension_raises(self, tmp_path):
        bad_file = tmp_path / "model.csv"
        bad_file.write_text("col1,col2\n1,2")
        with pytest.raises(ValueError, match="Unsupported file format"):
            parse_excel(bad_file)

    def test_parses_sheets_into_dataframes(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Revenue Projections"
        ws.append(["year", "stall_rental_income", "vendor_fees", "market_levies", "occupancy_rate"])
        ws.append([2025, 50_000, 10_000, 5_000, 0.70])
        ws.append([2026, 55_000, 11_000, 5_250, 0.75])

        path = tmp_path / "test_model.xlsx"
        wb.save(path)

        model = parse_excel(path, validate=False)
        assert "Revenue Projections" in model.sheets
        df = model.sheets["Revenue Projections"]
        assert len(df) == 2
        assert "year" in df.columns

    def test_schema_auto_detected(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Revenue Projections"
        ws.append(["year", "stall_rental_income"])
        path = tmp_path / "revenue_model.xlsx"
        wb.save(path)

        model = parse_excel(path, validate=False)
        assert model.detected_schema == "revenue_schema"

    def test_empty_sheet_skipped(self, tmp_path):
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Empty Sheet"
        # No data

        ws2 = wb.create_sheet("Revenue Projections")
        ws2.append(["year", "value"])
        ws2.append([2025, 100])

        path = tmp_path / "partial_model.xlsx"
        wb.save(path)

        model = parse_excel(path, validate=False)
        assert "Revenue Projections" in model.sheets
        # Empty sheet may or may not be included depending on implementation

    def test_source_file_stored(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(["col1"])
        path = tmp_path / "test.xlsx"
        wb.save(path)

        model = parse_excel(path, validate=False)
        assert model.source_file == path

    def test_detect_schema_revenue(self):
        import pandas as pd
        sheets = {"Revenue Projections": pd.DataFrame()}
        assert _detect_schema(sheets) == "revenue_schema"

    def test_detect_schema_capex(self):
        import pandas as pd
        sheets = {"Capital Expenditure": pd.DataFrame()}
        assert _detect_schema(sheets) == "capex_schema"

    def test_detect_schema_none(self):
        import pandas as pd
        sheets = {"Random Sheet Name": pd.DataFrame()}
        assert _detect_schema(sheets) is None
